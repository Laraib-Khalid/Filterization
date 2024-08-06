import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define file paths
file1 = 'D:/compare/data/excel/file1.xlsx'
file2 = 'D:/compare/data/excel/file2.xlsx'
output_file = 'D:/compare/data/excel/highlighted_output_processed.xlsx'

# Load Excel files into DataFrames
df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

# Define the column names to compare
columns = ['Price Rule', 'Terumo Product Code', 'Description', 'EA/BX', 'List Price (per pc)', 'List Price (per box)', 'Net Price (per pc)', 'Net Price (per box)']

# Check if all specified columns exist in the DataFrames
missing_columns_df1 = [col for col in columns if col not in df1.columns]
missing_columns_df2 = [col for col in columns if col not in df2.columns]

if missing_columns_df1:
    raise ValueError(f"Columns missing in file1: {', '.join(missing_columns_df1)}")
if missing_columns_df2:
    raise ValueError(f"Columns missing in file2: {', '.join(missing_columns_df2)}")

# Reorder columns in both DataFrames to match the specified order
df1 = df1[columns]
df2 = df2[columns]

# Create a DataFrame for df2 with the relevant columns, and convert it to a set of tuples for easy comparison
df2_tuples = set(df2.itertuples(index=False, name=None))

# Create a mask for matching rows in df1
def row_matches(row):
    return tuple(row) in df2_tuples

df1['Match'] = df1.apply(row_matches, axis=1)

# Check for duplicates in df1
df1['Duplicate'] = df1.duplicated(keep=False)

# Define the fill colors
highlight_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
highlight_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
highlight_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# Save df1 to a new Excel file with highlighting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df1.to_excel(writer, index=False, sheet_name='Sheet1')

    # Load the workbook and select the active sheet
    workbook = writer.book
    sheet = workbook['Sheet1']

    # Highlight rows based on conditions
    for row_idx, (match, duplicate) in enumerate(zip(df1['Match'], df1['Duplicate']), start=2):
        fill_color = None
        if match:
            if duplicate:
                fill_color = highlight_yellow
            else:
                fill_color = highlight_green
        else:
            fill_color = highlight_red

        if fill_color:
            for col_idx in range(1, len(df1.columns) + 1):  # Iterate through all columns
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color

print(f"Highlighted Excel file saved as {output_file}")
